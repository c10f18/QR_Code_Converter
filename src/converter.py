"""엑셀 읽기 → QR 생성/저장 핵심 로직 및 워커 스레드."""
import queue
import re
import threading
from pathlib import Path

import qrcode
from qrcode.constants import ERROR_CORRECT_M
from openpyxl import load_workbook

from logger import JobLogger

COL_URL = "sourceurl"
COL_NAME = "qrcodename"
_INVALID_FILENAME_CHARS = re.compile(r'[\\/:*?"<>|]')


def _normalize_header(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", "", str(value)).lower()


def read_rows(excel_path: str) -> list[tuple[str, str]]:
    """첫 시트에서 (Source URL, QR Code Name) 목록을 읽는다.

    헤더는 대소문자/공백 무관 매칭. 두 값이 모두 빈 행은 무시하고,
    한쪽만 빈 행은 그대로 반환하여 워커가 스킵 경고를 남기게 한다.
    """
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    try:
        ws = wb.worksheets[0]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            raise ValueError("엑셀 파일이 비어 있습니다.")

        col_url = col_name = None
        for idx, cell in enumerate(header):
            normalized = _normalize_header(cell)
            if normalized == COL_URL:
                col_url = idx
            elif normalized == COL_NAME:
                col_name = idx
        if col_url is None or col_name is None:
            raise ValueError(
                "필수 컬럼(Source URL, QR Code Name)을 찾을 수 없습니다.\n"
                "엑셀 템플릿 다운로드 기능으로 형식을 확인해 주세요."
            )

        rows: list[tuple[str, str]] = []
        for row in rows_iter:
            url = row[col_url] if col_url < len(row) else None
            name = row[col_name] if col_name < len(row) else None
            url = str(url).strip() if url is not None else ""
            name = str(name).strip() if name is not None else ""
            if not url and not name:
                continue  # 완전히 빈 행은 건수에서 제외
            rows.append((url, name))
        return rows
    finally:
        wb.close()


def sanitize_filename(name: str) -> str:
    """Windows에서 사용할 수 없는 문자를 _ 로 치환."""
    return _INVALID_FILENAME_CHARS.sub("_", name).strip().rstrip(".")


def unique_path(out_dir: Path, stem: str, ext: str) -> Path:
    """이미 같은 이름이 있으면 파일명(2), 파일명(3)… 으로 번호를 붙인다."""
    candidate = out_dir / f"{stem}{ext}"
    counter = 2
    while candidate.exists():
        candidate = out_dir / f"{stem}({counter}){ext}"
        counter += 1
    return candidate


class ConvertWorker(threading.Thread):
    """행 단위 QR 변환 워커.

    pause/resume: 닫기 확인 팝업이 떠 있는 동안 작업을 멈추기 위한 이벤트.
    cancel: '예' 선택 시 루프를 탈출하고 로그를 마무리한다.
    UI 갱신은 events 큐로만 전달한다(tkinter 위젯 직접 조작 금지).
    """

    def __init__(
        self,
        rows: list[tuple[str, str]],
        out_dir: Path,
        ext: str,
        box_size: int,
        border: int,
        events: queue.Queue,
        logger: JobLogger,
    ):
        super().__init__(daemon=True)
        self.rows = rows
        self.out_dir = Path(out_dir)
        self.ext = ext  # ".png" | ".jpg"
        self.box_size = box_size
        self.border = border
        self.events = events
        self.logger = logger
        self._pause = threading.Event()
        self._pause.set()
        self._cancel = threading.Event()

    def pause(self) -> None:
        self._pause.clear()

    def resume(self) -> None:
        self._pause.set()

    def cancel(self) -> None:
        self._cancel.set()
        self._pause.set()  # 일시정지 상태에서도 즉시 깨어나 종료 처리

    def _log(self, line: str) -> None:
        self.logger.write(line)
        self.events.put(("log", line))

    def _make_qr(self, url: str, name: str) -> Path:
        qr = qrcode.QRCode(
            error_correction=ERROR_CORRECT_M,
            box_size=self.box_size,
            border=self.border,
        )
        qr.add_data(url)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        pil = img.get_image() if hasattr(img, "get_image") else img

        stem = sanitize_filename(name) or "qr"
        path = unique_path(self.out_dir, stem, self.ext)
        if self.ext == ".jpg":
            pil = pil.convert("RGB")  # JPG는 1비트/투명 모드 미지원
        pil.save(path)
        return path

    def run(self) -> None:
        total = len(self.rows)
        success = 0
        fail = 0
        cancelled = False
        try:
            for n, (url, name) in enumerate(self.rows, start=1):
                self._pause.wait()
                if self._cancel.is_set():
                    cancelled = True
                    break

                if not url or not name:
                    fail += 1
                    self._log(
                        f"[{n}] Convert Skipped : 빈 값이 있는 행입니다 "
                        f"(Source URL='{url}', QR Code Name='{name}')"
                    )
                    self.events.put(("progress", n))
                    continue

                self._log(f"[{n}] Start convert URL to QR Code - {url} to {name}")
                try:
                    saved = self._make_qr(url, name)
                    success += 1
                    self._log(f"[{n}] Convert Success : {saved.name}")
                except Exception as e:  # 개별 실패는 기록 후 계속 진행
                    fail += 1
                    self._log(f"[{n}] Convert Failed : {name} - {e}")
                self.events.put(("progress", n))
        finally:
            self.logger.write("")
            if cancelled:
                self.logger.write("사용자 취소로 작업이 중단되었습니다.")
            self.logger.write(f"총 {total}건 중 {success}건 성공, {fail}건 실패")
            self.logger.close()
            self.events.put(
                (
                    "done",
                    {
                        "total": total,
                        "success": success,
                        "fail": fail,
                        "cancelled": cancelled,
                        "log_path": self.logger.path,
                    },
                )
            )
