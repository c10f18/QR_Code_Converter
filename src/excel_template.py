"""엑셀 템플릿(QR_Template.xlsx) 생성."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

HEADERS = ["Source URL", "QR Code Name"]
SAMPLE_ROWS = [
    ("https://www.example.com/event/1", "example_qr_01"),
    ("https://www.example.com/event/2", "example_qr_02"),
]
DEFAULT_FILENAME = "QR_Template.xlsx"


def create_template(path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "QR List"
    ws.append(HEADERS)
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for row in SAMPLE_ROWS:
        ws.append(row)
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 25
    wb.save(path)
